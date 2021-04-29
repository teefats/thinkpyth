from openpyxl import load_workbook
workbook = load_workbook(filename='Permit Tracker lite.xlsx')
names= workbook.sheetnames
print(names)